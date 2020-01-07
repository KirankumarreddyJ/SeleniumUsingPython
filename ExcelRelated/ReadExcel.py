import openpyxl

#To load the excel workbook
wb = openpyxl.load_workbook("../TestData/TestData.xlsx")

#To print the list of all sheet names
print(wb.sheetnames)

#To find the active sheet name
print("Active sheet name:" + wb.active.title)

#To select the sheet
ws = wb["Credentials"]
rows = ws.max_row
columns = ws.max_column
print('No.of rows : '+ str(rows) )
print('No.of columns : '+ str(columns))

#To print cell values
print(ws["A3"].value)
print(ws.cell(1,5).value)
print(ws.cell(column=6,row=1).value)

rows = ws.max_row
columns = ws.max_column
print('No.of rows : '+ str(rows) )
print('No.of columns : '+ str(columns))

print("******** For loop 1 ************* ")
for i in range(1,rows+1):
    for j in range(1, columns+1):
        cell = ws.cell(i,j)
        print(cell.value)
print("******** For loop 2 ************")
for r in ws['A1':'C4']:
    for c in r:
        print(c.value)
