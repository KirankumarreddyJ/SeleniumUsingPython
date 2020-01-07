import openpyxl

wk = openpyxl.Workbook()
ws = wk.active
print('active sheet name : ' + ws.title)

ws.title = "Data"
print(ws.title)

wk.create_sheet('NewSheet')

print(wk.worksheets)

sheet = wk['NewSheet']
ListVal = [10,20,30,40,50]
for i in range(1,6):
    sheet.cell(i,i).value = ListVal[i-1]

#Remove sheet from the excel
wk.remove(wk['Data'])
#To save the work book

wk.save('./Excel/WriteExcel4.xlsx')