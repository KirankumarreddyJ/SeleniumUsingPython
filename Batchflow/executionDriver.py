import openpyxl, os, sys, time, shutil
from FunctionLibrery.globalVariables import GlobalVariables as globVal
from TestScripts import *

dispatcher = {'TC_001_Login':TC_001_Login.TC_001_Login,'TC_002_Signup':TC_002_Signup.TC_002_Signup,
              'TC_004_WebTable':TC_004_WebTable.TC_004_WebTable,
              'TC_MC_001_LoginVerification':TC_MC_001_LoginVerification.TC_MC_001_LoginVerification}


def makeBatchResultDirectory():
    globVal.batchResultPath = "../Results/BatchResults/BatchResults_" + str(time.strftime("%Y%m%d-%H%M%S"))
    try:
        os.makedirs(globVal.batchResultPath)
    except Exception as e:
        print("Creation of the directory %s failed" % globVal.testResultFolderPath)
        raise Exception(e)

def copyFolder(src, dst, symlinks=False, ignore=None):
    for item in os.listdir(src):
        if str(item).lower() == str(globVal.testResultFolderName).lower():
            s = os.path.join(src, item)
            d = os.path.join(dst, item)
            if os.path.isdir(s):
                shutil.copytree(s, d, symlinks, ignore)
                break
            #else:
                #shutil.copy2(s, d)

def getScriptNamesToRunFromDataSheet():
    try:
        wb = openpyxl.load_workbook(filename=globVal.batchFlow_ExcelPath, read_only=True)
        sh = wb[globVal.batchFlow_TestCases_SheetName]
        scriptNamesToRun = []
        maxRow = sh.max_row
        maxColumn = sh.max_column
        scrColumNum = 0
        runStatusColumNum = 0
        for col in range(1, maxColumn + 1):
            if str(sh.cell(1, col).value).lower() == 'scriptname':
                scrColumNum = col
                if scrColumNum != 0 and runStatusColumNum != 0:
                    break
            elif str(sh.cell(1, col).value).lower() == 'runstatus':
                runStatusColumNum = col
                if scrColumNum != 0 and runStatusColumNum != 0:
                    break
        else:
            if scrColumNum == 0 and runStatusColumNum == 0:
                print('Both "ScriptName" and "RunStatus" columns are not present under sheet('+globVal.batchFlow_TestCases_SheetName+') in "'+str(os.path.abspath(globVal.batchFlow_ExcelPath))+'"')
                sys.exit(1)
            elif scrColumNum == 0 :
                print('"ScriptName" column is not present under sheet('+globVal.batchFlow_TestCases_SheetName+') in "'+str(os.path.abspath(globVal.batchFlow_ExcelPath))+'"')
                sys.exit(1)
            elif runStatusColumNum == 0:
                print('"RunStatus" column is not present under sheet(' + globVal.batchFlow_TestCases_SheetName + ') in "' + str(os.path.abspath(globVal.batchFlow_ExcelPath)) + '"')
                sys.exit(1)
        for r in range(2, maxRow +1):
            if str(sh.cell(r,runStatusColumNum).value).lower() == 'yes':
                scriptName = str(sh.cell(r,scrColumNum).value)
                scriptNamesToRun.append([r,scriptName])
        wb._archive.close()
        return scriptNamesToRun
    except Exception as e:
        print('Got Exception in "getScriptNamesToRunFromDataSheet" function. Exception is :' + str(e))
        raise Exception(e)

def writeResultToDataSheet(filePath, scriptRowNum, testResultStatus):
    try:
        wb = openpyxl.load_workbook(filename=filePath)
        sh = wb[globVal.batchFlow_TestCases_SheetName]

        maxRow = sh.max_row
        maxColumn = sh.max_column
        resStatusColumNum = 0
        resLinkColumNum = 0
        for col in range(1, maxColumn + 1):
            val = str(sh.cell(1, col).value).lower()
            if val == 'resultstatus':
                resStatusColumNum = col
                if resStatusColumNum != 0 and resLinkColumNum != 0:
                    break
            elif val == 'resultlink':
                resLinkColumNum = col
                if resStatusColumNum != 0 and resLinkColumNum != 0:
                    break
        else:
            if resStatusColumNum == 0 and resLinkColumNum == 0:
                print('Both "testResultStatus" and "ResultLink" columns are not present under sheet('+globVal.batchFlow_TestCases_SheetName+') in "'+str(os.path.abspath(globVal.batchFlow_ExcelPath))+'"')
                sys.exit(1)
            elif resStatusColumNum == 0 :
                print('"testResultStatus" column is not present under sheet('+globVal.batchFlow_TestCases_SheetName+') in "'+str(os.path.abspath(globVal.batchFlow_ExcelPath))+'"')
                sys.exit(1)
            elif resLinkColumNum == 0:
                print('"ResultLink" column is not present under sheet(' + globVal.batchFlow_TestCases_SheetName + ') in "' + str(os.path.abspath(globVal.batchFlow_ExcelPath)) + '"')
                sys.exit(1)

        sh.cell(int(scriptRowNum),resStatusColumNum).value = testResultStatus
        sh.cell(int(scriptRowNum),resLinkColumNum).value = '=HYPERLINK("{}", "{}")'.format(os.path.abspath(str(globVal.testResultFolderPath)) + "/Results.html", "Result.html")
        wb.save(os.path.abspath(filePath))

    except Exception as excp:
        print('Got Excepiton in "writeResultToDataSheet" function. Exception is :' +str(excp))
        raise Exception(excp)

def removeAllResultStatusesFromDataSheet(filePath):
    try:
        wb = openpyxl.load_workbook(filename=filePath)
        sh = wb[globVal.batchFlow_TestCases_SheetName]
        maxRow = sh.max_row
        maxColumn = sh.max_column
        runStatusColumNum = 0
        resStatusColumNum = 0
        resLinkColumNum = 0
        for col in range(1, maxColumn + 1):
            if str(sh.cell(1, col).value).lower() == 'runstatus':
                runStatusColumNum = col
                if runStatusColumNum != 0 and resStatusColumNum != 0 and resLinkColumNum != 0:
                    break
            elif str(sh.cell(1, col).value).lower() == 'resultstatus':
                resStatusColumNum = col
                if runStatusColumNum != 0 and resStatusColumNum != 0 and resLinkColumNum != 0:
                    break
            elif str(sh.cell(1, col).value).lower() == 'resultlink':
                resLinkColumNum = col
                if runStatusColumNum != 0 and resStatusColumNum != 0 and resLinkColumNum != 0:
                    break
        else:
            if runStatusColumNum == 0 and resStatusColumNum == 0 and resLinkColumNum == 0:
                print("While removing the result status from batchflow sheet",'Both "RunStatus", "ResultStatus" and "ResultLink" columns are not present under sheet('+globVal.batchFlow_TestCases_SheetName+') in "'+str(os.path.abspath(globVal.batchFlow_ExcelPath))+'"')
            elif runStatusColumNum == 0:
                print("While removing the result status from batchflow sheet",'"RunStatus" column is not present under sheet(' + globVal.batchFlow_TestCases_SheetName + ') in "' + str(os.path.abspath(globVal.batchFlow_ExcelPath)) + '"')
            elif resStatusColumNum == 0:
                print("While removing the result status from batchflow sheet",'"ResultStatus" column is not present under sheet(' + globVal.batchFlow_TestCases_SheetName + ') in "' + str(os.path.abspath(globVal.batchFlow_ExcelPath)) + '"')
            elif resLinkColumNum == 0:
                print("While removing the result status from batchflow sheet",'"ResultLink" column is not present under sheet(' + globVal.batchFlow_TestCases_SheetName + ') in "' + str(os.path.abspath(globVal.batchFlow_ExcelPath)) + '"')
        for r in range(2, maxRow + 1):
            if str(sh.cell(r, runStatusColumNum).value).lower() == 'yes':
                sh.cell(r, resStatusColumNum).value = ''
                sh.cell(r, resLinkColumNum).value = ''
        wb.save(os.path.abspath(filePath))

    except Exception as e:
        print('Got Exception in funtion(removeAllResultStatusesFromDataSheet)', 'Exception is : '+str(e))

def executeBatch():
    try:
        scriptNamesToRun = getScriptNamesToRunFromDataSheet()
        makeBatchResultDirectory()
        globVal.batchResultExcelPath = shutil.copy(str(globVal.batchFlow_ExcelPath), str(os.path.abspath(globVal.batchResultPath))+'/BatchResults.xlsx')
        totalScriptsSelected = len(scriptNamesToRun)
        totalScriptsExecuted = 0
        totalScriptsPassed = 0
        totalScriptsFailed = 0
        totalScriptsWarning = 0
        totalScriptsStopped = 0
        totalScriptsWithUnknowStatus = 0
        print()
        print('****************************************************************')
        print('Batch Execution Started, Find below batch execution details:')
        print('Total Number Of Scripts Selected = '+ str(totalScriptsSelected))
        print('Names Of Selected Scripts = ', end = '')
        for i in range(len(scriptNamesToRun)):
            if i != len(scriptNamesToRun)-1:
                print(str(scriptNamesToRun[i][1]), end=', ')
            else:
                print(str(scriptNamesToRun[i][1]))
        print('****************************************************************')

        for a in scriptNamesToRun:
            scriptRowNum = int(a[0])
            scriptFunction = dispatcher[a[1]]
            print('----------------------------------------------')
            print('Started Executing Script: "' + str(a[1])+'",', 'Started Time: "'+ str(time.strftime("%Y%m%d-%H%M%S"))+'"')
            scriptFunction()
            totalScriptsExecuted += 1
            ResultStatus=''
            if globVal.totalFailed >0:
                ResultStatus = 'Failed'
                totalScriptsFailed += 1
            elif globVal.totalWarning >0:
                ResultStatus = 'Warning'
                totalScriptsWarning += 1
            elif str(globVal.scriptCompletelyExecuted).lower() != 'yes':
                ResultStatus = 'Stopped'
                totalScriptsStopped += 1
            elif str(globVal.scriptCompletelyExecuted).lower() == 'yes':
                ResultStatus = 'Passed'
                totalScriptsPassed += 1
            else:
                ResultStatus = 'Something Went Wrong'
                totalScriptsWithUnknowStatus += 1
            print('Execution Completed. Test Status: "' + str(ResultStatus)+'",', 'Test EndTime: "'+ str(time.strftime("%Y%m%d-%H%M%S"))+'"')
            print('----------------------------------------------')
            writeResultToDataSheet(globVal.batchFlow_ExcelPath,scriptRowNum,ResultStatus)
            writeResultToDataSheet(globVal.batchResultExcelPath,scriptRowNum,ResultStatus)
            sourceFolder = "../Results/TestResults/"
            destinationFolder = str(os.path.abspath(str(globVal.batchResultPath)))
            copyFolder(sourceFolder, destinationFolder)
            globVal.resetGlobalVals()

        removeAllResultStatusesFromDataSheet(globVal.batchFlow_ExcelPath)
        print('****************************************************************')
        print("Batch Execution Is Completed. Below are the batch details:")
        print("Total Scripts Selected: " + str(totalScriptsSelected))
        print("Total Scripts Executed: " + str(totalScriptsExecuted))
        print("Total Scripts Passed: "+ str(totalScriptsPassed))
        print("Total Scripts Failed: "+ str(totalScriptsFailed))
        print("Total Scripts Warnings: " + str(totalScriptsWarning))
        print("Total Scripts Incompletely Executed: " + str(totalScriptsStopped))
        if totalScriptsWithUnknowStatus > 0:
            print("Total Scripts With Unknown Status: '" +str(totalScriptsWithUnknowStatus)+"'")
        print('****************************************************************')

    except Exception as e:
        print('Got exception. Error is : "' + str(e) + '"')
        raise Exception('Got exception. Error is : "' + str(e) + '"')

if __name__ == "__main__":
    executeBatch()
